"""
Inventory Dashboard Builder
Generates inventory-dashboard.xlsx with three sections mirroring the Google Sheets dashboard:
  A:B  — Inventory Pivot (all non-zero pick SKUs, sorted by available qty)
  E:G  — Issue SKUs (pick SKUs where ending balance < 0, with their Shopify SKU mapping)
  L:O  — Inventory Transaction / Committed (ending balance if staged orders are committed)

Run this script whenever you want to refresh the dashboard from updated source sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling helpers ────────────────────────────────────────────────────────────

FONT_NAME = "Arial"

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Color palette
C_HEADER_BG    = "2E75B6"   # dark blue — section headers
C_PICK_BG      = "D6E4F0"   # light blue — pick SKU group rows
C_TOTAL_BG     = "BDD7EE"   # medium blue — group total rows
C_GRAND_BG     = "1F4E79"   # darkest blue — grand total row
C_NEG_BG       = "FFE0E0"   # light red — negative ending balance
C_POS_BG       = "E2EFDA"   # light green — positive/OK
C_SHOPIFY_BG   = "FFFFFF"   # white — shopify SKU detail rows
C_ALT_BG       = "F2F2F2"   # light gray — alternate row tint for inventory pivot


def style_cell(cell, value=None, bold=False, font_color="000000", bg=None,
               h_align="left", number_format=None, italic=False):
    if value is not None:
        cell.value = value
    cell.font = font(bold=bold, color=font_color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.border = border()
    cell.alignment = align(h=h_align, v="center")
    if number_format:
        cell.number_format = number_format


# ── Source data ────────────────────────────────────────────────────────────────

INVENTORY = [
    # Positive (have stock) — listed first
    ("apple_pink_lady-01x02",    236),
    ("avocaodo_hass-48x25",      140),
    ("apple_cosmic-01x02",        80),
    ("apple_envy-01x02",          70),
    ("starfruit-01x04",           11),
    ("kumkquat-BG01x02",          10),
    ("df_red-10x07",               8),
    # Negative (overcommitted)
    ("pf_purple-01x12",         -287),
    ("mandarin_satsuma-01x04",  -373),
    ("mango_cherry-09x07",      -461),
    ("cherimoya-W01x01",        -603),
    ("pf_purple-BG0101",        -629),
    ("blood_orange-01x04",      -682),
    ("avocado_hass-48x25",      -734),
    ("mango_honey-09x12",       -740),
    ("plum_sugar-BG01x01",      -744),
    ("sapote_white-01x03",      -779),
    ("apple_honeycrisp-01x02",  -804),
    ("honey-04x16",             -845),
    ("date_dry_barhi-BG01x01",  -851),
    ("loquat-BG01x01",          -852),
    ("date_medjool-BG0101",     -917),
    ("pear_asian-12x18",        -918),
    ("cherry-01x01",            -924),
    ("pineapple_pink-02x01",    -942),
    ("df_yellow-05x07",         -987),
    ("tangerine_sumo-07x10",    -991),
    ("df_white-10x07",          -994),
]

# pick_sku → [shopify_skus] that map to it (from SKU mapping table)
SKU_MAPPING = {
    "apple_honeycrisp-01x02": [
        "f.apple_honeycrisp-2lb",
        "f.apple_honeycrisp-2lb-bab",
    ],
    "avocado_hass-48x25": [
        "f.avocado_hass-2lb",
        "f.avocado_reed-2lb",
        "f.avocado_variety-2lb",
        "m.exoticfruit_beginner-3lb",
    ],
    "blood_orange-01x04": [
        "f.blood_orange-0.5lb-gift",
        "f.orange_blood-2lb-bab",
        "m.exoticfruit_beginner-3lb",
        "m.exoticfruit-8lb",
    ],
    "cherimoya-W01x01": [
        "f.cherimoya-2lb",
        "m.exoticfruit-3lb-bab",
        "m.exoticfruit-8lb",
    ],
    "cherry-01x01": ["f.cherry_bing-2lb-bab"],
    "date_medjool-BG0101": ["f.dates_medjool-2lb"],
    "df_white-10x07": ["f.dragonfruit_variety-2lb-bab"],
    "df_yellow-05x07": [
        "f.dragonfruit_variety-2lb-bab",
        "f.dragonfruit_yellow-2lb",
        "m.exoticfruit-8lb",
    ],
    "honey-04x16": ["f.raw_honey-9oz-addon"],
    "loquat-BG01x01": [
        "f.loquat-2lb",
        "f.loquat-5lb",
        "f.loquat-8lb",
        "m.exoticfruit-3lb-bab",
        "m.exoticfruit-8lb",
    ],
    "mandarin_satsuma-01x04": [
        "f.mandarin_satsuma-2lb",
        "f.nc.gift-B",
    ],
    "mango_cherry-09x07": [
        "f.mango_cherry-2lb",
        "f.mango_cherry-2lb-bab",
        "f.mango_cherry-5lb",
        "f.mango-0.5lb-gift",
        "m.exoticfruit_beginner-3lb",
        "m.exoticfruit-3lb-bab",
        "m.exoticfruit-8lb",
    ],
    "mango_honey-09x12": [
        "f.mango_honey-2lb-bab",
        "m.exoticfruit-8lb",
    ],
    "pear_asian-12x18": [
        "f.asian_pear-2lb",
        "f.asian_pear-5lb",
    ],
    "pf_purple-01x12": [
        "f.nc.gift-B",
        "f.passion_fruit-6pc-gift",
        "m.exoticfruit_beginner-3lb",
        "m.exoticfruit-3lb-bab",
        "m.exoticfruit-8lb",
    ],
    "pf_purple-BG0101": ["f.passionfruit_purple-2lb"],
    "pineapple_pink-02x01": ["f.pineapple_pink-2lb"],
    "plum_sugar-BG01x01": [
        "f.cherry-plum-2lb-bab",
        "f.plum_sugar-2lb-bab",
        "m.exoticfruit-8lb",
    ],
    "sapote_white-01x03": [
        "f.sapote_white-2lb",
        "m.exoticfruit_beginner-3lb",
        "m.exoticfruit-8lb",
    ],
    "tangerine_sumo-07x10": ["f.tangerine_sumo-2lb-bab"],
}

# Staged demand: pick_sku → total pick units being deducted by staged orders
STAGED_DEMAND = {
    "starfruit-01x04":         4,
    "kumkquat-BG01x02":        3,
    "pf_purple-01x12":        19,
    "mandarin_satsuma-01x04": 31,
    "mango_cherry-09x07":     40,
    "cherimoya-W01x01":       18,
    "pf_purple-BG0101":       10,
    "blood_orange-01x04":     35,
    "plum_sugar-BG01x01":     18,
    "mango_honey-09x12":      23,
    "avocado_hass-48x25":     37,
    "sapote_white-01x03":      9,
    "apple_honeycrisp-01x02":  8,
    "honey-04x16":             1,
    "date_medjool-BG0101":     2,
    "cherry-01x01":            2,
    "pear_asian-12x18":       11,
    "pineapple_pink-02x01":    3,
    "loquat-BG01x01":        106,
    "df_yellow-05x07":         8,
    "df_white-10x07":          3,
    "tangerine_sumo-07x10":    9,
}


# ── Derived data ───────────────────────────────────────────────────────────────

inv_lookup = dict(INVENTORY)

# Section C: all pick_skus with staged demand, sorted by ending balance asc
section_c_rows = []
for sku, demand_qty in STAGED_DEMAND.items():
    avail = inv_lookup.get(sku, 0)
    ending = avail - demand_qty
    section_c_rows.append((sku, demand_qty, ending, -demand_qty))
section_c_rows.sort(key=lambda x: x[2])  # most negative first

grand_qty      = sum(r[1] for r in section_c_rows)
grand_ending   = min(r[2] for r in section_c_rows)
grand_deduct   = -grand_qty

# Section B: pick_skus where ending balance < 0
issue_rows = [(sku, shopify_skus, inv_lookup.get(sku, 0) - STAGED_DEMAND[sku])
              for sku, shopify_skus in SKU_MAPPING.items()
              if sku in STAGED_DEMAND and inv_lookup.get(sku, 0) - STAGED_DEMAND[sku] < 0]
issue_rows.sort(key=lambda x: x[2])  # most negative first


# ── Build workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
dash = wb.active
dash.title = "Dashboard"
dash.sheet_view.showGridLines = True

# Freeze top 2 rows (section label + column headers)
dash.freeze_panes = "A3"

# ── Row 1: Section banners ─────────────────────────────────────────────────────
BANNER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)

def banner(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BANNER_FONT
    cell.fill = fill(C_HEADER_BG)
    cell.alignment = align(h="center", v="center")
    cell.border = border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

banner(dash, 1, 1,  "Inventory Pivot",                 2)   # A:B
banner(dash, 1, 5,  "Issue SKU",                        3)   # E:G
banner(dash, 1, 12, "Inventory Transaction / Committed", 4)  # L:O

# ── Row 2: Column headers ──────────────────────────────────────────────────────
COL_HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)

def col_header(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = COL_HEADER_FONT
    cell.fill = fill("4472C4")
    cell.alignment = align(h="center", v="center", wrap=True)
    cell.border = border()

col_header(dash, 2, 1,  "Pick SKU")
col_header(dash, 2, 2,  "Available Qty")
col_header(dash, 2, 5,  "Pick SKU")
col_header(dash, 2, 6,  "Shopify SKU")
col_header(dash, 2, 7,  "Ending Balance")
col_header(dash, 2, 12, "Pick SKU")
col_header(dash, 2, 13, "Staged Qty")
col_header(dash, 2, 14, "Ending Balance")
col_header(dash, 2, 15, "Qty Deduction")

# ── Column widths ──────────────────────────────────────────────────────────────
dash.column_dimensions["A"].width = 30
dash.column_dimensions["B"].width = 14
dash.column_dimensions["C"].width = 3   # spacer
dash.column_dimensions["D"].width = 3   # spacer
dash.column_dimensions["E"].width = 28
dash.column_dimensions["F"].width = 30
dash.column_dimensions["G"].width = 14
dash.column_dimensions["H"].width = 3   # spacer
dash.column_dimensions["I"].width = 3
dash.column_dimensions["J"].width = 3
dash.column_dimensions["K"].width = 3   # spacer
dash.column_dimensions["L"].width = 28
dash.column_dimensions["M"].width = 12
dash.column_dimensions["N"].width = 14
dash.column_dimensions["O"].width = 14

# Row heights
dash.row_dimensions[1].height = 20
dash.row_dimensions[2].height = 30

# ── Section A: Inventory Pivot (cols 1:2) ─────────────────────────────────────
NUM_FMT = '#,##0;(#,##0);"-"'

row = 3
for i, (sku, avail) in enumerate(INVENTORY):
    bg = C_POS_BG if avail > 0 else C_NEG_BG
    alt_bg = "F2F9F0" if avail > 0 else "FFF5F5"
    row_bg = bg if i % 2 == 0 else alt_bg

    cell_sku = dash.cell(row=row, column=1, value=sku)
    cell_sku.font = font(bold=(avail > 0))
    cell_sku.fill = fill(row_bg)
    cell_sku.border = border()
    cell_sku.alignment = align()

    cell_qty = dash.cell(row=row, column=2, value=avail)
    cell_qty.font = font(bold=(avail > 0))
    cell_qty.fill = fill(row_bg)
    cell_qty.border = border()
    cell_qty.alignment = align(h="right")
    cell_qty.number_format = NUM_FMT
    row += 1

# Grand total — Section A
style_cell(dash.cell(row=row, column=1), "Grand Total",
           bold=True, font_color="FFFFFF", bg=C_GRAND_BG, h_align="center")
gt_val = sum(q for _, q in INVENTORY)
cell = dash.cell(row=row, column=2, value=gt_val)
cell.font = font(bold=True, color="FFFFFF")
cell.fill = fill(C_GRAND_BG)
cell.border = border()
cell.alignment = align(h="right")
cell.number_format = NUM_FMT

# ── Section B: Issue SKUs (cols 5:7) ──────────────────────────────────────────
row = 3
for pick_sku, shopify_skus, ending_bal in issue_rows:
    # Pick SKU header row
    style_cell(dash.cell(row=row, column=5), pick_sku,
               bold=True, bg=C_PICK_BG, h_align="left")
    dash.cell(row=row, column=6).fill = fill(C_PICK_BG)
    dash.cell(row=row, column=6).border = border()
    style_cell(dash.cell(row=row, column=7), ending_bal,
               bold=True, bg=C_PICK_BG, h_align="right", number_format=NUM_FMT)
    row += 1

    # Shopify SKU detail rows
    for s_sku in shopify_skus:
        dash.cell(row=row, column=5).fill = fill(C_SHOPIFY_BG)
        dash.cell(row=row, column=5).border = border()

        style_cell(dash.cell(row=row, column=6), s_sku,
                   italic=True, bg=C_SHOPIFY_BG)

        style_cell(dash.cell(row=row, column=7), ending_bal,
                   bg=C_NEG_BG, h_align="right", number_format=NUM_FMT)
        row += 1

    # Group total row
    style_cell(dash.cell(row=row, column=5), f"{pick_sku} Total",
               bold=True, bg=C_TOTAL_BG, h_align="right")
    dash.cell(row=row, column=6).fill = fill(C_TOTAL_BG)
    dash.cell(row=row, column=6).border = border()
    style_cell(dash.cell(row=row, column=7), ending_bal,
               bold=True, bg=C_TOTAL_BG, h_align="right", number_format=NUM_FMT)
    row += 1

# Grand total — Section B
style_cell(dash.cell(row=row, column=5), "Grand Total",
           bold=True, font_color="FFFFFF", bg=C_GRAND_BG, h_align="center")
dash.cell(row=row, column=6).fill = fill(C_GRAND_BG)
dash.cell(row=row, column=6).border = border()
style_cell(dash.cell(row=row, column=7), min(e for _, _, e in issue_rows),
           bold=True, font_color="FFFFFF", bg=C_GRAND_BG, h_align="right",
           number_format=NUM_FMT)

# ── Section C: Inventory Transaction / Committed (cols 12:15) ─────────────────
row = 3
for i, (sku, staged_qty, ending, deduction) in enumerate(section_c_rows):
    row_bg = C_NEG_BG if ending < 0 else C_POS_BG
    alt_bg = "FFF5F5" if ending < 0 else "F2F9F0"
    actual_bg = row_bg if i % 2 == 0 else alt_bg

    style_cell(dash.cell(row=row, column=12), sku, bg=actual_bg)
    style_cell(dash.cell(row=row, column=13), staged_qty,
               bg=actual_bg, h_align="right", number_format=NUM_FMT)
    style_cell(dash.cell(row=row, column=14), ending,
               bold=(ending < 0), bg=actual_bg, h_align="right", number_format=NUM_FMT)
    style_cell(dash.cell(row=row, column=15), deduction,
               bg=actual_bg, h_align="right", number_format=NUM_FMT)
    row += 1

# Grand total — Section C
for col, val in [(12, "Grand Total"), (13, grand_qty),
                 (14, grand_ending), (15, grand_deduct)]:
    h = "center" if col == 12 else "right"
    cell = dash.cell(row=row, column=col, value=val)
    cell.font = font(bold=True, color="FFFFFF")
    cell.fill = fill(C_GRAND_BG)
    cell.border = border()
    cell.alignment = align(h=h, v="center")
    if col != 12:
        cell.number_format = NUM_FMT


# ── Source sheets (for future data refresh) ────────────────────────────────────

# Sheet: Inventory Source
inv_ws = wb.create_sheet("Inventory Source")
for col, hdr in enumerate(["pick_sku", "available_qty"], 1):
    c = inv_ws.cell(row=1, column=col, value=hdr)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill("2E75B6")
    c.border = border()
    c.alignment = align(h="center")
inv_ws.column_dimensions["A"].width = 32
inv_ws.column_dimensions["B"].width = 14
for r, (sku, qty) in enumerate(INVENTORY, 2):
    inv_ws.cell(row=r, column=1, value=sku).border = border()
    c = inv_ws.cell(row=r, column=2, value=qty)
    c.border = border()
    c.number_format = NUM_FMT
    c.alignment = align(h="right")

# Sheet: Staged Demand Source
dem_ws = wb.create_sheet("Staged Demand Source")
for col, hdr in enumerate(["pick_sku", "staged_qty"], 1):
    c = dem_ws.cell(row=1, column=col, value=hdr)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill("2E75B6")
    c.border = border()
    c.alignment = align(h="center")
dem_ws.column_dimensions["A"].width = 32
dem_ws.column_dimensions["B"].width = 12
for r, (sku, qty) in enumerate(STAGED_DEMAND.items(), 2):
    dem_ws.cell(row=r, column=1, value=sku).border = border()
    c = dem_ws.cell(row=r, column=2, value=qty)
    c.border = border()
    c.number_format = NUM_FMT
    c.alignment = align(h="right")

# Sheet: SKU Mapping Source
map_ws = wb.create_sheet("SKU Mapping Source")
for col, hdr in enumerate(["pick_sku", "shopify_sku"], 1):
    c = map_ws.cell(row=1, column=col, value=hdr)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill("2E75B6")
    c.border = border()
    c.alignment = align(h="center")
map_ws.column_dimensions["A"].width = 32
map_ws.column_dimensions["B"].width = 32
map_r = 2
for pick_sku, shopify_list in sorted(SKU_MAPPING.items()):
    for s_sku in shopify_list:
        map_ws.cell(row=map_r, column=1, value=pick_sku).border = border()
        map_ws.cell(row=map_r, column=2, value=s_sku).border = border()
        map_r += 1

# ── Save ───────────────────────────────────────────────────────────────────────
OUTPUT = "/Users/robertfan/Claude Code/fulfillment-app/data/inventory-dashboard.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
